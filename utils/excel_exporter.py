import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(bold=True, color="E94560", size=11)
ALT_FILL = PatternFill("solid", fgColor="16213E")
NORMAL_FONT = Font(color="FFFFFF", size=10)
BORDER = Border(
    bottom=Side(style="thin", color="E94560"),
)

def _style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER

def _style_row(cell, alt=False):
    cell.fill = PatternFill("solid", fgColor="0F3460" if alt else "16213E")
    cell.font = NORMAL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")

def _write_table(ws, headers, rows, start_row=1):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        _style_header(cell)
    
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _style_row(cell, alt=(row_idx % 2 == 0))
    
    # Auto width
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22


def generate_excel(stock_data: dict, order_data: dict) -> bytes:
    wb = openpyxl.Workbook()
    
    # ---- Sheet 1: Stock Details ----
    ws1 = wb.active
    ws1.title = "Stock Details"
    ws1.sheet_view.showGridLines = False
    
    headers1 = ["Field", "Value"]
    
    rows1 = [
        ["Script Name", stock_data.get("script_name", "N/A")],
        ["Instrument Type", stock_data.get("instrument_type", "equity").capitalize()],
        ["LTP (Last Traded Price)", stock_data.get("ltp", "N/A")],
        ["Open Price", stock_data.get("open", "N/A")],
        ["Previous Close", stock_data.get("close", "N/A")],
        ["Day High", stock_data.get("day_high", "N/A")],
        ["Day Low", stock_data.get("day_low", "N/A")],
        ["Upper Circuit", stock_data.get("upper_circuit", "N/A")],
        ["Lower Circuit", stock_data.get("lower_circuit", "N/A")],
        ["52-Week High", stock_data.get("week_52_high", "N/A")],
        ["52-Week Low", stock_data.get("week_52_low", "N/A")],
    ]
    
    if stock_data.get("instrument_type") == "futures":
        rows1.append(["Expiry Date", stock_data.get("expiry_date", "N/A")])
        rows1.append(["Lot Size", stock_data.get("lot_size", "N/A")])
    
    if stock_data.get("index_note"):
        rows1.append(["Note", stock_data.get("index_note")])
    
    _write_table(ws1, headers1, rows1)
    
    # ---- Sheet 2: Order Summary ----
    ws2 = wb.create_sheet("Order Summary")
    ws2.sheet_view.showGridLines = False
    
    headers2 = ["Field", "Value"]
    rows2 = [
        ["Product Type", order_data.get("product_type", "N/A")],
        ["Order Type", order_data.get("order_type", "N/A")],
        ["Quantity / Lots", order_data.get("quantity", "N/A")],
        ["LTP Used", order_data.get("ltp", "N/A")],
        ["Estimated Amount Required (₹)", order_data.get("estimated_amount", "N/A")],
    ]
    
    _write_table(ws2, headers2, rows2)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
